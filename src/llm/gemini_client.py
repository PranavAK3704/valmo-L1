"""
Gemini client using the new google-genai SDK.
Returns structured JSON decisions from the model.
"""

import json
import logging
import os
import re
import time

from google import genai
from google.genai import types

logger = logging.getLogger(__name__)

MODEL = "gemini-3-flash-preview"

# Gemini 2.5 Flash pricing (USD per 1M tokens)
PRICE_INPUT_PER_M  = 0.15   # $0.15 / 1M input tokens
PRICE_OUTPUT_PER_M = 0.60   # $0.60 / 1M output tokens
USD_TO_INR         = 84.0   # approximate

def tokens_to_inr(input_tokens: int, output_tokens: int) -> float:
    cost_usd = (input_tokens * PRICE_INPUT_PER_M + output_tokens * PRICE_OUTPUT_PER_M) / 1_000_000
    return round(cost_usd * USD_TO_INR, 4)

SYSTEM_PROMPT = """You are a Valmo L1 support agent. Your job is to resolve tickets raised by delivery partners (captains).

You think and act exactly like a trained human L1 agent:
- You read the ticket carefully
- If a STAGE 0 — SITUATION ASSESSMENT block is present, treat it as a pre-reasoned diagnosis from the supply chain analyzer. Use it as your starting point. Override it only when query data clearly contradicts.
- You check what the SOP says for this type of issue
- You look at the Metabase/Log10 query data provided
- You apply the SOP logic to reach a decision
- If you are genuinely unsure or the data is insufficient, you say so honestly

## Output Format
Always respond with ONLY a valid JSON object — no markdown, no extra text:

{
  "action": "respond" | "escalate" | "needs_info" | "stuck",
  "clean_problem": "...",         // one sentence: what the captain needs in plain ops language (e.g. "Hardstop loss on AWB VL123 — captain claims valid delivery within SLA — needs Log10 scan verification")
  "response_to_captain": "...",   // draft reply to captain (for action=respond or needs_info)
  "escalation_queue": "...",      // e.g. L2_LOSSES, L2_PAYMENTS, L2_OPERATIONS (for action=escalate)
  "escalation_reason": "...",     // why escalating (for action=escalate)
  "stuck_question": "...",        // your specific question for the trainer (for action=stuck)
  "missing_fields": ["..."],      // what info to request from captain (for action=needs_info)
  "confidence": 0-10,             // how confident you are (>=7 = auto-send, <7 = needs review)
  "scenario_identified": "...",   // EXACT scenario ID from the STRUCTURED SOP DECISION TREE (e.g. HS_5, HS_WAIVER, SS_3, CON_PAY_1) — NEVER use broad category names like "hardstop_loss"
  "reasoning": "..."              // brief explanation of your decision
}

## Rules
- NEVER make up data. If query results are empty or missing, say so.
- NEVER fill in amounts, dates, or AWBs that aren't in the provided data.
- **CAPTAIN'S CLAIMS ARE NOT VERIFIED DATA.** When the captain writes things like "I checked Metabase and it shows X" or "Log10 says Y" or "Recovered from Pilot per LM FE Loss Marked" — that is the CAPTAIN'S claim, not evidence. The agent's value is doing independent verification. If the SOP scenario requires checking a specific Metabase query or Log10 scan and the METABASE QUERY RESULTS section below is empty or doesn't include that query, you MUST use action=stuck (not escalate, not respond) and your stuck_question must specify exactly which query needs to be run. Do not escalate-for-reversal based on the captain's own claim of having checked the data.
- If the SOP says escalate to L2 or L3, always escalate — do not try to resolve it yourself.
- If you don't have enough data to decide, use action=stuck and ask a specific question.
- Confidence < 7 means the response goes to trainer for review before sending.
- Write responses in a professional but friendly tone — same as a human support agent.
- VAGUE TICKETS: If the captain uses undefined terms like "suspicious shipment", "some issue", "problem with parcel" without explaining WHAT the suspicion or problem actually is — do NOT guess a scenario. Use action=needs_info and ask specifically what the suspicion/issue is. Example: "Could you please clarify what was suspicious about the shipment — was it tampered packaging, a customer fraud concern, incorrect contents, or something else?"
- READ ALL SCENARIOS in the SOP, not just the first one that matches the keywords. A ticket about "not receiving evidence mail" maps to Scenario 1 (mails not sent to partner = L1→L3 escalation), NOT Scenario 5 (wrong email ID shared). Read carefully before deciding.
- If ANY scenario in the SOP says "direct L1 to L3 escalation", use action=escalate with escalation_queue=L3.
- CRITICAL — "delivered" in ALL loss/hardstop scenarios means delivered to the END CUSTOMER (last-mile: the FE/driver physically handed the package to the customer). It does NOT mean delivered to the hub. "Shipment delivered but couldn't be marked" (Scenario 7) = FE went to customer, gave the package, but forgot to mark it in the app. If a ticket says a product is MISSING from a box that arrived at the hub, that is a MISSING ITEM / SHORTAGE event — apply the shortage SOP, NOT Scenario 7.
- If the ticket mentions proof/attachments are shared (CCTV, photos, weight slips, etc.), note this in your reasoning and factor it into the decision (e.g., if proof is shared, escalate with proof noted; if no proof, request it).
"""


def _extract_json(text: str) -> dict:
    """Extract JSON from model response, handling markdown code blocks."""
    text = re.sub(r"```(?:json)?\s*", "", text).strip()
    text = text.rstrip("`").strip()
    return json.loads(text)


class GeminiClient:
    def __init__(self):
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY not set in environment")
        self._client = genai.Client(api_key=api_key)
        logger.info(f"[Gemini] Initialized with model={MODEL}")

    def decide(self, ticket_context: str, retries: int = 3) -> dict:
        """
        Send ticket context to Gemini and get a structured decision back.
        Returns parsed dict with action, response, confidence, etc.
        """
        for attempt in range(retries):
            try:
                response = self._client.models.generate_content(
                    model=MODEL,
                    contents=ticket_context,
                    config=types.GenerateContentConfig(
                        system_instruction=SYSTEM_PROMPT,
                        temperature=0.1,
                    ),
                )
                raw = response.text.strip()
                logger.debug(f"[Gemini] Raw response: {raw[:500]}")
                result = _extract_json(raw)
                # Attach token usage to result
                try:
                    um = response.usage_metadata
                    inp = um.prompt_token_count or 0
                    out = um.candidates_token_count or 0
                    result["_usage"] = {
                        "input_tokens": inp,
                        "output_tokens": out,
                        "cost_inr": tokens_to_inr(inp, out),
                    }
                except Exception:
                    result["_usage"] = {"input_tokens": 0, "output_tokens": 0, "cost_inr": 0.0}
                return result
            except json.JSONDecodeError as e:
                logger.warning(f"[Gemini] JSON parse failed (attempt {attempt+1}): {e}")
                if attempt == retries - 1:
                    return {
                        "action": "stuck",
                        "stuck_question": f"Model returned unparseable response: {raw[:200]}",
                        "confidence": 0,
                        "reasoning": "JSON parse error",
                        "scenario_identified": "unknown",
                        "_usage": {"input_tokens": 0, "output_tokens": 0, "cost_inr": 0.0},
                    }
            except Exception as e:
                logger.error(f"[Gemini] API error (attempt {attempt+1}): {e}")
                if attempt < retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    return {
                        "action": "stuck",
                        "stuck_question": f"Gemini API error: {e}",
                        "confidence": 0,
                        "reasoning": "API error",
                        "scenario_identified": "unknown",
                        "_usage": {"input_tokens": 0, "output_tokens": 0, "cost_inr": 0.0},
                    }
        return {"action": "stuck", "confidence": 0, "reasoning": "exhausted retries", "scenario_identified": "unknown",
                "_usage": {"input_tokens": 0, "output_tokens": 0, "cost_inr": 0.0}}

    def generate_json(self, prompt: str, system_prompt: str = "", temperature: float = 0.1,
                      expect_list: bool = False, retries: int = 2):
        """
        Generic JSON generation call — used by comprehension & matching stages.
        Returns parsed dict or list depending on expect_list.
        """
        for attempt in range(retries):
            try:
                cfg = types.GenerateContentConfig(temperature=temperature)
                if system_prompt:
                    cfg = types.GenerateContentConfig(
                        system_instruction=system_prompt, temperature=temperature
                    )
                response = self._client.models.generate_content(
                    model=MODEL, contents=prompt, config=cfg,
                )
                raw = response.text.strip()
                raw = re.sub(r"```(?:json)?\s*", "", raw).strip().rstrip("`").strip()
                result = json.loads(raw)
                # Attach usage if caller wants it (stored as _usage on dicts, ignored on lists)
                try:
                    um = response.usage_metadata
                    inp = um.prompt_token_count or 0
                    out = um.candidates_token_count or 0
                    usage = {"input_tokens": inp, "output_tokens": out, "cost_inr": tokens_to_inr(inp, out)}
                    if isinstance(result, dict):
                        result["_usage"] = usage
                    else:
                        result = {"_list": result, "_usage": usage}
                except Exception:
                    pass
                return result
            except Exception as e:
                logger.warning(f"[Gemini.generate_json] attempt {attempt+1} failed: {e}")
                if attempt < retries - 1:
                    time.sleep(1)
        return [] if expect_list else {}


    def _read_tabular_local(self, path: str, ext: str) -> str:
        """Parse .xlsx/.xls/.csv into a compact text dump suitable for both
        AWB-regex scanning and inclusion in the LLM prompt. Returns rows as
        '| cell1 | cell2 | ...' lines, sheet headers if multiple sheets.
        Raises if the file can't be parsed."""
        import csv as _csv

        if ext == ".csv":
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                reader = _csv.reader(f)
                rows = [r for r in reader if any(c.strip() for c in r)]
            return self._dump_rows(rows, sheet_name="(csv)")

        # .xlsx / .xls — use openpyxl. For old .xls we may need xlrd; if not
        # installed, the openpyxl call will raise and the caller falls through
        # to Gemini Vision.
        try:
            import openpyxl
        except Exception as e:
            raise RuntimeError(f"openpyxl not installed: {e}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() for c in row):
                    rows.append([("" if c is None else str(c)) for c in row])
            if rows:
                parts.append(self._dump_rows(rows, sheet_name=sheet_name))
        wb.close()
        return "\n\n".join(parts) if parts else "(empty workbook)"

    def _dump_rows(self, rows: list, sheet_name: str = "") -> str:
        """Format a list-of-lists as a markdown-ish table for LLM consumption."""
        if not rows:
            return ""
        head = f"### Sheet: {sheet_name}\n" if sheet_name else ""
        # Cap to 200 rows so a 10k-row dump doesn't blow the prompt budget
        truncated = len(rows) > 200
        rows_out = rows[:200]
        lines = ["| " + " | ".join((c or "").strip() for c in r) + " |" for r in rows_out]
        body = "\n".join(lines)
        if truncated:
            body += f"\n…({len(rows) - 200} more rows truncated)"
        return head + body

    def read_attachment(self, url: str, filename: str = "", prompt: str = None) -> str:
        """
        Download a PDF/image/Excel attachment from an S3 URL and extract text.

        Routing:
          * .xlsx / .xls / .csv  → parsed LOCALLY (openpyxl/csv). Free, fast, no
                                   API call. Returns the cell contents as
                                   structured text so AWB-regex can scan it.
          * .pdf / images        → Gemini Vision via Files API.
          * Anything else        → "[Attachment type ... not supported]".

        `prompt` only matters for the Gemini-Vision path. Excel/CSV are always
        dumped as raw cell text regardless of prompt.
        """
        import urllib.request, tempfile, pathlib, csv as _csv

        ext = pathlib.Path(filename or url.split("?")[0]).suffix.lower()
        IMAGE_OR_PDF = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp"}
        TABULAR      = {".xlsx", ".xls", ".csv"}
        if ext not in (IMAGE_OR_PDF | TABULAR):
            return f"[Attachment type {ext} not supported for auto-reading]"

        # Download once into a temp file
        try:
            with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                tmp_path = tmp.name
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                with open(tmp_path, "wb") as f:
                    f.write(resp.read())
        except Exception as e:
            logger.warning(f"[Gemini.read_attachment] download failed for {filename or url[:60]}: {e}")
            return f"[Could not read attachment: download failed: {e}]"

        # ── Tabular: parse locally, no Gemini call ──
        if ext in TABULAR:
            try:
                text = self._read_tabular_local(tmp_path, ext)
                return text[:6000]   # bigger budget than vision OCR since AWB lists can be long
            except Exception as e:
                logger.warning(f"[Gemini.read_attachment] local parse failed for {filename}: {e}")
                # Fall through to Gemini Vision as last resort
            finally:
                try: pathlib.Path(tmp_path).unlink()
                except Exception: pass

        # ── PDF / image: Gemini Vision ──
        default_prompt = (
            "Extract all text, AWB numbers (format: VL followed by digits, or similar tracking IDs), "
            "shipment IDs, amounts, dates, and any structured data from this document. "
            "List all AWB/tracking numbers clearly. If it is a loss statement or shortage report, "
            "extract: DC code, date range, shipment count, total loss amount, and each AWB listed."
        )
        try:
            upload = self._client.files.upload(file=tmp_path)
            response = self._client.models.generate_content(
                model=MODEL,
                contents=[upload, prompt or default_prompt],
                config=types.GenerateContentConfig(temperature=0.0),
            )
            try: self._client.files.delete(name=upload.name)
            except Exception: pass
            try: pathlib.Path(tmp_path).unlink()
            except Exception: pass

            return response.text.strip()[:3000]

        except Exception as e:
            logger.warning(f"[Gemini.read_attachment] Failed to read {filename or url[:60]}: {e}")
            return f"[Could not read attachment: {e}]"


# Singleton
_client: GeminiClient | None = None

def get_gemini_client() -> GeminiClient:
    global _client
    if _client is None:
        _client = GeminiClient()
    return _client
